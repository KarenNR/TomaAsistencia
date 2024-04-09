from flask import Flask, render_template, request, redirect, session, flash
from flask_mysqldb import MySQL
from datetime import datetime, timedelta
import time
import openpyxl
import os

# ---------- GLOBAL VARIABLES ----------

semesterStart = datetime(2024, 1, 10)
semesterEnd = datetime(2024, 5, 10)
todayDate = datetime.now()
holidays = (datetime(2024, 2, 5), datetime(2024, 3, 18), datetime(2024, 3, 25), 
            datetime(2024, 3, 26), datetime(2024, 3, 27), datetime(2024, 3, 28), 
            datetime(2024, 3, 29), datetime(2024, 3, 30), datetime(2024, 5, 1))

# ---------- FLASK CONFIGURATION ----------

app = Flask(__name__)
app.secret_key = "dae2024"

app.config['MYSQL_HOST'] = 'SECRET'
app.config['MYSQL_USER'] = 'SECRET'
app.config['MYSQL_PASSWORD'] = 'SECRET'
app.config['MYSQL_DB'] = 'SECRET'

mysql = MySQL(app)

# ---------- WEB ROUTES ----------

@app.route('/')
def loadIndex():
    return render_template("index.html")


@app.route('/login', methods=['POST'])
def login():
    _username = request.form['username']
    _password = request.form['password']
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT Nomina, Nombre, Apellido FROM Profesor WHERE Usuario=(%s) AND Pass=(%s)''', (_username, _password))
    data = cursor.fetchone()
    if data is not None:
        session['nomina'] = data[0]
        session['profesor'] = data[1] + ' ' + data[2]
        return redirect('/cursos')
    else:
        flash('El usuario o la contraseña no coinciden.')
        return redirect('/')


@app.route('/cursos')
def loadMenu():
    if 'profesor' in session:
        cursor = mysql.connection.cursor()
        cursor.execute('''SELECT * FROM Clase WHERE Nomina_Profesor=(%s)''', (session['nomina'],))
        courses = cursor.fetchall()
        # Format hour
        x = []
        for element in courses:
            y = list(element)
            y[10] = ":".join(str(element[10]).split(":")[0:2])
            y[11] = ":".join(str(element[11]).split(":")[0:2])
            element = tuple(y)
            x.append(element)
        courses = tuple(x)
        cursor.close()
        return render_template("profesor/class-menu.html", profesor=session['profesor'], cursos=courses)
    else:
        flash("Por favor inicie sesión antes de ingresar al sistema")
        return redirect("/")


@app.route('/informacion-curso/<int:id>')
def loadCourseInformation(id):
    if 'profesor' in session:
        # Get course information
        cursor = mysql.connection.cursor()
        cursor.execute('''SELECT * FROM Clase WHERE ID_Clase=(%s)''', (id,))
        courseInformation = cursor.fetchone()
        # Format hour
        x = list(courseInformation)
        x[10] = ":".join(str(courseInformation[10]).split(":")[0:2])
        x[11] = ":".join(str(courseInformation[11]).split(":")[0:2])
        courseInformation = tuple(x)
        cursor.close()
        # Get days in which class is taken
        days = courseInformation[3:10]
        # Get all students from course
        cursor = mysql.connection.cursor()
        cursor.execute('''
                        SELECT Matricula, Alumno.Nombre, Apellido, Carrera 
                        FROM Clase JOIN Alumno_Clase ON Clase.ID_Clase=Alumno_Clase.ID_Clase
                            JOIN Alumno On Matricula=Matricula_Alumno
                        WHERE Clase.ID_Clase=(%s)
                        ORDER BY Matricula''', (id,))
        students = cursor.fetchall()
        cursor.close()
        # Calculate attendance average per student
        classAverage = {"asistencia": 0, "falta": 0, "retardo": 0, "total": 0}
        studentAverages = []
        for student in students:
            attendanceInfo = getStudentAverage(id, student[0], days)
            studentAverages.append(attendanceInfo)
            classAverage["asistencia"] += attendanceInfo["asistencia"]
            classAverage["falta"] += attendanceInfo["falta"]
            classAverage["retardo"] += attendanceInfo["retardo"]
            classAverage["total"] += attendanceInfo["total"]

        return render_template("profesor/informacion-curso.html", profesor=session['profesor'], curso=courseInformation, 
                            alumnos=zip(students, studentAverages), promedioClase=classAverage, cantidadAlumnos=len(students))
    else:
        flash("Por favor inicie sesión antes de ingresar al sistema")
        return redirect("/")


@app.route('/informacion-curso/<int:courseId>/alumno/<int:studentId>')
def loadStudentInformation(courseId, studentId):
    if 'profesor' in session:
        # Get course information
        cursor = mysql.connection.cursor()
        cursor.execute('''SELECT * FROM Clase WHERE ID_Clase=(%s)''', (courseId,))
        courseInformation = cursor.fetchone()
        cursor.close()
        # Get days in which class is taken
        days = courseInformation[3:10]
        # Get student information
        cursor = mysql.connection.cursor()
        cursor.execute('''SELECT * FROM Alumno WHERE Matricula=(%s)''', (studentId,))
        studentInformation = cursor.fetchone()
        cursor.close()
        # Calculate attendance average
        average, details = getStudentDetail(courseId, studentId, days)
        return render_template("profesor/informacion-alumno.html", profesor=session['profesor'], alumno=studentInformation,
                            promedio=average, detalles=details, curso=courseId, nombreCurso=courseInformation[1])
    else:
        flash("Por favor inicie sesión antes de ingresar al sistema")
        return redirect("/")


# ---------- DOWNLOAD REPORTS ROUTES ---------

@app.route('/descargar-reporte/<int:courseId>/alumno/<int:studentId>')
def downloadStudentReporte(courseId, studentId):
    # Get course information
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * FROM Clase WHERE ID_Clase=(%s)''', (courseId,))
    courseInformation = cursor.fetchone()
    cursor.close()
    # Get days in which class is taken
    days = courseInformation[3:10]
    # Get student information
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * FROM Alumno WHERE Matricula=(%s)''', (studentId,))
    studentInformation = cursor.fetchone()
    cursor.close()
    # Calculate attendance average
    average, details = getStudentDetail(courseId, studentId, days)
    # Create XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data = [
        ["REPORTE DE ASISTENCIAS DE ALUMNO"],
        ["Fecha", todayDate.strftime("%d-%m-%Y")],
        ["Clase", courseInformation[1]],
        [],
        ["DATOS DEL ALUMNO"],
        ["Matricula", "Nombre", "Apellidos", "Carrera"],
        list(studentInformation),
        [],
        ["PROMEDIO DE ASISTENCIAS"],
        ["Asistencias", "Faltas", "Retardos", "Promedio al día de hoy"],
        [average["asistencia"], average["falta"], average["retardo"], average["asistencia"]/average["total"]],
        [],
        ["ASISTENCIAS POR DÍA"]
    ]
    for detail in details:
        content = [detail[0], "Asistencia" if detail[1]==0 else "Falta" if detail[1]==1 else "Retardo"]
        data.append(content)
    for row in data:
        sheet.append(row)
    # Format data
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)
    sheet.merge_cells(start_row=13, start_column=1, end_row=13, end_column=2)
    column_widths = [20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    sheet['D11'].number_format = '0.00%'
    sheet['A1'].font = openpyxl.styles.Font(bold=True)
    sheet['A5'].font = openpyxl.styles.Font(bold=True)
    sheet['A9'].font = openpyxl.styles.Font(bold=True)
    sheet['A13'].font = openpyxl.styles.Font(bold=True)
    # Download report
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    file_name = downloads_path + "/Reporte_Alumno_{}_{}.xlsx".format(studentId, int(time.time()))
    workbook.save(file_name)
    # Send success message
    flash("Reporte guardado en descargas")
    return redirect("/informacion-curso/{}/alumno/{}".format(courseId, studentId))


@app.route('/descargar-reporte/<int:id>/<string:type>')
def downloadTodayReport(id, type):
    reportDate = todayDate.strftime("%Y-%m-%d") if type == "today" else type
    # Get course information
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * FROM Clase WHERE ID_Clase=(%s)''', (id,))
    courseInformation = cursor.fetchone()
    cursor.close()
    # Get all students from course
    cursor = mysql.connection.cursor()
    cursor.execute('''
                    SELECT Matricula, Alumno.Nombre, Apellido, Carrera 
                    FROM Clase JOIN Alumno_Clase ON Clase.ID_Clase=Alumno_Clase.ID_Clase
                        JOIN Alumno On Matricula=Matricula_Alumno
                    WHERE Clase.ID_Clase=(%s)
                    ORDER BY Matricula''', (id,))
    students = cursor.fetchall()
    cursor.close()
    # Check attendance per student
    classAverage = {"asistencia": 0, "falta": 0, "retardo": 0, "total": 0}
    studentAttendance = []
    for student in students:
        row = [student[0], student[1], student[2]]
        cursor = mysql.connection.cursor()
        cursor.execute('''SELECT Asistencia FROM Asistencia WHERE ID_Clase=(%s) AND Matricula_Alumno=(%s) AND Fecha=(%s)''',
                       (id, student[0], reportDate))
        attendance = cursor.fetchone()
        # 0 = Asistencia / 1 = Falta / 2 = Retardo
        if attendance is None:
            classAverage["falta"] += 1
            row.append(1)
        elif attendance[0] == 0:
            classAverage["asistencia"] += 1
            row.append(0)
        elif attendance[0] == 1:
            classAverage["falta"] += 1
            row.append(1)
        else:
            classAverage["retardo"] += 1
            row.append(2)
        classAverage["total"] += 1
        studentAttendance.append(row)
    # Create XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data = [
        ["REPORTE DE ASISTENCIAS POR DÍA"],
        ["Fecha", reportDate],
        ["Clase", courseInformation[1]],
        [],
        ["PROMEDIO DE ASISTENCIAS"],
        ["Asistencias", "Faltas", "Retardos", "Promedio del día"],
        [classAverage["asistencia"], classAverage["falta"], classAverage["retardo"], classAverage["asistencia"]/classAverage["total"]],
        [],
        ["DESGLOSE DE ASISTENCIA"],
        ["Matrícula", "Nombre", "Apellidos", "Asistencia"]
    ]
    for detail in studentAttendance:
        content = [detail[0], detail[1], detail[2], "Asistencia" if detail[3]==0 else "Falta" if detail[3]==1 else "Retardo"]
        data.append(content)
    for row in data:
        sheet.append(row)
    # Format data
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)
    column_widths = [20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    sheet['D7'].number_format = '0.00%'
    sheet['A1'].font = openpyxl.styles.Font(bold=True)
    sheet['A5'].font = openpyxl.styles.Font(bold=True)
    sheet['A9'].font = openpyxl.styles.Font(bold=True)
    # Download report
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    file_name = downloads_path + "/Reporte_Clase_{}_{}_{}.xlsx".format(id, reportDate, int(time.time()))
    workbook.save(file_name)
    # Send success message
    flash("Reporte guardado en descargas")
    return redirect("/informacion-curso/{}".format(id))


@app.route('/descargar-reporte/<int:id>/general')
def downloadGeneralReport(id):
    # Get course information
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * FROM Clase WHERE ID_Clase=(%s)''', (id,))
    courseInformation = cursor.fetchone()
    cursor.close()
    # Get days in which class is taken
    days = courseInformation[3:10]
    # Get all students from course
    cursor = mysql.connection.cursor()
    cursor.execute('''
                    SELECT Matricula, Alumno.Nombre, Apellido, Carrera 
                    FROM Clase JOIN Alumno_Clase ON Clase.ID_Clase=Alumno_Clase.ID_Clase
                        JOIN Alumno On Matricula=Matricula_Alumno
                    WHERE Clase.ID_Clase=(%s)
                    ORDER BY Matricula''', (id,))
    students = cursor.fetchall()
    cursor.close()
    # Get average attendance per student
    classAverage = {"asistencia": 0, "falta": 0, "retardo": 0, "total": 0}
    studentAttendance = []
    for student in students:
        attendanceInfo = getStudentAverage(id, student[0], days)
        classAverage["asistencia"] += attendanceInfo["asistencia"]
        classAverage["falta"] += attendanceInfo["falta"]
        classAverage["retardo"] += attendanceInfo["retardo"]
        classAverage["total"] += attendanceInfo["total"]
        row = [student[0], student[1], student[2], attendanceInfo["asistencia"]/attendanceInfo["total"]]
        studentAttendance.append(row)
    # Create XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data = [
        ["REPORTE DE PROMEDIO GRUPAL DE ASISTENCIAS"],
        ["Fecha", todayDate.strftime("%d-%m-%Y")],
        ["Clase", courseInformation[1]],
        [],
        ["PROMEDIO GRUPAL DE ASISTENCIAS"],
        ["Asistencias", "Faltas", "Retardos", "Promedio al día de hoy"],
        [classAverage["asistencia"], classAverage["falta"], classAverage["retardo"], classAverage["asistencia"]/classAverage["total"]],
        [],
        ["PROMEDIO DE ASISTENCIA POR ALUMNO"],
        ["Matrícula", "Nombre", "Apellidos", "Promedio de asistencia"]
    ]
    for detail in studentAttendance:
        data.append(detail)
    for row in data:
        sheet.append(row)
    # Format data
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)
    column_widths = [20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    for cell in sheet['D']:
        cell.number_format = '0.00%'
    sheet['A1'].font = openpyxl.styles.Font(bold=True)
    sheet['A5'].font = openpyxl.styles.Font(bold=True)
    sheet['A9'].font = openpyxl.styles.Font(bold=True)
    # Download report
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    file_name = downloads_path + "/Reporte_Promedio_Clase_{}_{}.xlsx".format(id, int(time.time()))
    workbook.save(file_name)
    # Send success message
    flash("Reporte guardado en descargas")
    return redirect("/informacion-curso/{}".format(id))


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


# ---------- ATTENDANCE AVERAGE FUNCTIONS ----------

def getStudentAverage(courseId, studentId, days):
    startDate = semesterStart
    endDate = todayDate if todayDate < semesterEnd else semesterEnd
    results = {"asistencia": 0, "falta": 0, "retardo": 0, "total": 0}
    while startDate <= endDate:
        if days[startDate.weekday()] and startDate not in holidays:
            cursor = mysql.connection.cursor()
            cursor.execute('''SELECT Asistencia FROM Asistencia WHERE ID_Clase=(%s) AND Matricula_Alumno=(%s) AND Fecha=(%s)''',
                           (courseId, studentId, startDate.strftime("%Y-%m-%d")))
            attendance = cursor.fetchone()
            # 0 = Asistencia / 1 = Falta / 2 = Retardo
            if attendance is None:
                results["falta"] += 1
            elif attendance[0] == 0:
                results["asistencia"] += 1
            elif attendance[0] == 1:
                results["falta"] += 1
            else:
                results["retardo"] += 1
            results["total"] += 1
        startDate += timedelta(days=1)
    return results


def getStudentDetail(courseId, studentId, days):
    startDate = semesterStart
    endDate = todayDate if todayDate < semesterEnd else semesterEnd
    results = {"asistencia": 0, "falta": 0, "retardo": 0, "total": 0}
    details = []
    while startDate <= endDate:
        if days[startDate.weekday()] and startDate not in holidays:
            detail = [startDate.strftime("%d-%m-%Y")]
            cursor = mysql.connection.cursor()
            cursor.execute('''SELECT Asistencia FROM Asistencia WHERE ID_Clase=(%s) AND Matricula_Alumno=(%s) AND Fecha=(%s)''',
                           (courseId, studentId, startDate.strftime("%Y-%m-%d")))
            attendance = cursor.fetchone()
            # 0 = Asistencia / 1 = Falta / 2 = Retardo
            if attendance is None:
                results["falta"] += 1
                detail.append(1)
            elif attendance[0] == 0:
                results["asistencia"] += 1
                detail.append(0)
            elif attendance[0] == 1:
                results["falta"] += 1
                detail.append(1)
            else:
                results["retardo"] += 1
                detail.append(2)
            results["total"] += 1
            details.append(detail)
        startDate += timedelta(days=1)
    return results, details
